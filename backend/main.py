from flask import Flask, jsonify, request, send_file
from flask_cors import CORS
import requests
from typing import List, Dict, Any
import os
import tempfile
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from dotenv import load_dotenv
import logging
import sys
import os
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from router_service import RouterService

load_dotenv()

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = Flask(__name__, static_folder='.', static_url_path='')
CORS(app)

# Initialize router service
router_service = RouterService()

# Rootstock APIs
ROOTSTOCK_API_BASE = "https://rootstock.blockscout.com/api/v2"
ROOTSTOCK_EXPLORER_API = "https://be.explorer.rootstock.io/api/v3"

class TokenBalance:
    def __init__(self, data):
        self.token = data.get("token", {})
        self.value = data.get("value", "0")

    def to_dict(self):
        return {
            "token": {
                "address_hash": self.token.get("address_hash"),
                "decimals": self.token.get("decimals"),
                "exchange_rate": self.token.get("exchange_rate"),
                "icon_url": self.token.get("icon_url"),
                "name": self.token.get("name"),
                "symbol": self.token.get("symbol"),
                "total_supply": self.token.get("total_supply"),
                "type": self.token.get("type"),
            },
            "token_id": self.token.get("token_id"),
            "token_instance": self.token.get("token_instance"),
            "value": self.value,
        }

@app.route("/")
def root():
    """Serve the main frontend page"""
    try:
        return app.send_static_file('index.html')
    except Exception as e:
        return f"Error serving frontend: {str(e)}", 500

@app.route("/health")
def health_check():
    return jsonify({"status": "healthy"})

@app.route("/api/token-balances/<address>")
def get_token_balances(address: str):
    """
    Fetch token balances for a given address from Rootstock Blockscout API
    """
    try:
        url = f"{ROOTSTOCK_API_BASE}/addresses/{address}/token-balances"
        response = requests.get(url, timeout=10)
        
        if response.status_code != 200:
            return jsonify({
                "error": f"Failed to fetch token balances: {response.text}"
            }), response.status_code
        
        data = response.json()
        
        # Process and format the data, excluding ERC-721 tokens (NFTs)
        token_balances = []
        for item in data:
            # Skip ERC-721 tokens as they are handled by the NFT service
            if item.get("token", {}).get("type") == "ERC-721":
                continue
                
            token_balance = TokenBalance(item)
            token_balances.append(token_balance.to_dict())
        
        return jsonify({
            "address": address,
            "token_balances": token_balances,
            "total_tokens": len(token_balances)
        })
        
    except requests.RequestException as e:
        return jsonify({"error": f"Request error: {str(e)}"}), 500
    except Exception as e:
        return jsonify({"error": f"Internal server error: {str(e)}"}), 500

def get_native_rbtc_balance(address):
    """
    Fetch native rBTC balance from Rootstock explorer API
    """
    try:
        # Explorer API requires lowercase addresses
        lowercase_address = address.lower()
        url = f"{ROOTSTOCK_EXPLORER_API}/balances/address/{lowercase_address}?take=1"
        
        response = requests.get(url, timeout=10)
        
        if response.status_code != 200:
            return None
        
        data = response.json()
        
        if "data" in data and len(data["data"]) > 0:
            # Get the latest balance (first item in the list)
            latest_balance = data["data"][0]
            balance_str = latest_balance.get("balance", "0")
            
            # Balance is already formatted as decimal string in v3 API
            balance_rbtc = float(balance_str)
            
            return {
                "balance": balance_str,
                "balance_formatted": balance_rbtc,
                "symbol": "rBTC",
                "name": "Rootstock Smart Bitcoin",
                "decimals": "18",
                "is_native": True
            }
        
        return None
        
    except Exception as e:
        return None


@app.route("/api/address-info/<address>")
def get_address_info(address: str):
    """
    Get basic address information and token balances
    """
    try:
        # Get token balances directly from Blockscout API
        url = f"{ROOTSTOCK_API_BASE}/addresses/{address}/token-balances"
        response = requests.get(url, timeout=10)
        
        if response.status_code != 200:
            return jsonify({
                "error": f"Failed to fetch token balances: {response.text}"
            }), response.status_code
        
        data = response.json()
        
        # Pre-process data for router service (expects list of dicts with token and value keys)
        router_balances = []
        for item in data:
            router_balances.append({
                "token": {
                    "type": item.get("token", {}).get("type"),
                    "symbol": item.get("token", {}).get("symbol"),
                    "name": item.get("token", {}).get("name"),
                    "address_hash": item.get("token", {}).get("address_hash")
                },
                "value": item.get("value", "0")
            })
        
        # Process through router service (optimized - only runs needed services)
        results = router_service.process_address(address, router_balances)
        
        # Get native rBTC balance from Explorer
        native_rbtc = get_native_rbtc_balance(address)
        
        # Process and format the data for response (exclude ERC-721 tokens to avoid duplication with NFT service)
        all_balances = []
        for item in data:
            # Skip ERC-721 tokens as they are handled by the NFT service
            if item.get("token", {}).get("type") == "ERC-721":
                continue
            token_balance = TokenBalance(item)
            all_balances.append(token_balance.to_dict())
        
        if native_rbtc:
            # Add native rBTC to the beginning of the list
            native_token_data = {
                "token": {
                    "address_hash": "0x0000000000000000000000000000000000000000",  # Native token address
                    "decimals": "0",  # Native rBTC balance is already in decimal format
                    "exchange_rate": None,  # Will be set from RBTC price
                    "icon_url": "https://assets.coingecko.com/coins/images/5070/small/RBTC-logo.png?1718152038",
                    "name": "Rootstock Smart Bitcoin",
                    "symbol": "rBTC",
                    "type": "native",
                    "volume_24h": None
                },
                "token_id": None,
                "token_instance": None,
                "value": native_rbtc["balance"]
            }
            
            # Find RBTC or WRBTC price from existing tokens to use for native rBTC
            rbtc_price = None
            for balance in all_balances:
                if balance["token"]["symbol"] in ["RBTC", "WRBTC"]:
                    rbtc_price = balance["token"]["exchange_rate"]
                    break
            
            # If no RBTC/WRBTC price found in existing tokens, try to get WRBTC price from Blockscout API
            if not rbtc_price:
                try:
                    # Get WRBTC price from Blockscout API using the WRBTC contract address - as the API has no price for the native asset. 
                    wrbtc_address = "0x542fda317318ebf1d3deaf76e0b632741a7e677d"
                    wrbtc_url = f"https://rootstock.blockscout.com/api/v2/tokens/{wrbtc_address}"
                    
                    logger.info(f"Fetching WRBTC price from: {wrbtc_url}")
                    response = requests.get(wrbtc_url, timeout=10)
                    logger.info(f"Blockscout API response status: {response.status_code}")
                    
                    if response.status_code == 200:
                        data = response.json()
                        logger.info(f"Blockscout API response: {data}")
                        if "exchange_rate" in data and data["exchange_rate"]:
                            rbtc_price = float(data["exchange_rate"])
                            logger.info(f"Retrieved WRBTC price from Blockscout: ${rbtc_price}")
                        else:
                            logger.warning(f"No exchange_rate found in Blockscout response: {data}")
                    else:
                        logger.warning(f"Blockscout API returned status {response.status_code}: {response.text}")
                except Exception as e:
                    logger.warning(f"Could not fetch WRBTC price from Blockscout API: {str(e)}")
            
            if rbtc_price:
                native_token_data["token"]["exchange_rate"] = rbtc_price
            
            # Rename RBTC to Wrapped Rootstock Smart Bitcoin
            for balance in all_balances:
                if balance["token"]["symbol"] == "RBTC":
                    balance["token"]["name"] = "Wrapped Rootstock Smart Bitcoin"
                    balance["token"]["symbol"] = "WRBTC"
                    break
            
            all_balances.insert(0, native_token_data)
        
        return jsonify({
            "address": address,
            "total_value_usd": len(all_balances),  # Placeholder
            "token_count": len(all_balances),
            "token_balances": all_balances,
            "nft_valuations": results["nft_valuations"],
            "nft_count": len(results["nft_valuations"]),
            "nft_total_value_usd": sum(v["total_value_usd"] for v in results["nft_valuations"]),
            "merkle_rewards": results["merkle_rewards"]["rewards"],
            "merkle_rewards_count": results["merkle_rewards"]["total_rewards"],
            "merkle_rewards_total_usd": results["merkle_rewards"]["total_usd_value"],
            "yield_tokens": results["yield_tokens"]["yield_tokens"],
            "yield_tokens_count": results["yield_tokens"]["total_yield_tokens"],
            "lending_portfolio": results["lending_portfolio"],
            "evidence": results["evidence"]  # Include evidence for debugging
        })
        
    except Exception as e:
        return jsonify({"error": f"Error fetching address info: {str(e)}"}), 500

# Yield tokens endpoint removed - now handled by /api/address-info/<address>

# Lending data endpoint removed - now handled by /api/address-info/<address>

# Tropykus portfolio endpoint removed - now handled by /api/address-info/<address>

@app.route("/api/export-excel", methods=['POST'])
def export_to_excel():
    """
    Export portfolio data to Excel spreadsheet using cached data from frontend
    """
    try:
        # Get data from request body (sent by frontend)
        data = request.get_json()
        
        if not data:
            return jsonify({"error": "No data provided"}), 400
        
        # Create Excel workbook
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create sheets using the same data structure as frontend
        create_wallet_sheet(wb, data.get("token_balances", []))
        create_portfolio_sheet(wb, data)
        create_summary_sheet(wb, data.get("address", "Unknown"), data)
        
        # Save to temporary file
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        wb.save(temp_file.name)
        temp_file.close()
        
        # Generate filename with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        address = data.get("address", "Unknown")
        filename = f"portfolio_{address[:8]}_{timestamp}.xlsx"
        
        return send_file(
            temp_file.name,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        return jsonify({"error": f"Error creating Excel file: {str(e)}"}), 500

def create_wallet_sheet(wb, token_balances):
    """Create Wallet sheet with token balances"""
    ws = wb.create_sheet("Wallet")
    
    # Headers
    headers = ["Token", "Name", "Holdings", "Price", "USD Value"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Data
    for row, token_data in enumerate(token_balances, 2):
        token = token_data["token"]
        value = token_data["value"]
        
        # Calculate values
        balance = float(value) / (10 ** int(token.get("decimals", 18)))
        price = float(token.get("exchange_rate", 0)) if token.get("exchange_rate") else 0
        usd_value = balance * price
        
        ws.cell(row=row, column=1, value=token.get("symbol", ""))
        ws.cell(row=row, column=2, value=token.get("name", ""))
        ws.cell(row=row, column=3, value=round(balance, 8))
        ws.cell(row=row, column=4, value=f"${price:,.2f}" if price > 0 else "N/A")
        ws.cell(row=row, column=5, value=f"${usd_value:,.2f}" if usd_value > 0 else "N/A")
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

def create_portfolio_sheet(wb, data):
    """Create Portfolio sheet using the same data structure as frontend"""
    ws = wb.create_sheet("Portfolio")
    
    # Headers
    headers = ["Type", "Protocol", "Name", "Holdings", "Price", "APR", "USD Value"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    row = 2
    
    # Process all portfolio items from the data structure
    # This mirrors how the frontend processes the data
    
    # Yield tokens
    yield_tokens = data.get("yield_tokens", [])
    if isinstance(yield_tokens, dict):
        yield_tokens = yield_tokens.get("yield_tokens", [])
    
    for yield_token in yield_tokens:
        ws.cell(row=row, column=1, value="Yield Token")
        ws.cell(row=row, column=2, value=yield_token.get("protocol", "Yield Protocol"))
        ws.cell(row=row, column=3, value=yield_token.get("name", "Unknown"))
        ws.cell(row=row, column=4, value=round(float(yield_token.get("balance", 0)), 8))
        ws.cell(row=row, column=5, value=f"${float(yield_token.get('price', 0)):,.2f}")
        ws.cell(row=row, column=6, value=f"{float(yield_token.get('apr', 0)):.2f}%")
        ws.cell(row=row, column=7, value=f"${float(yield_token.get('usd_value', 0)):,.2f}")
        row += 1
    
    # Lending positions - process all protocols generically
    lending_portfolio = data.get("lending_portfolio", {})
    for protocol_name, protocol_data in lending_portfolio.items():
        if isinstance(protocol_data, dict) and "portfolio_items" in protocol_data:
            # Tropykus format
            for item in protocol_data["portfolio_items"]:
                ws.cell(row=row, column=1, value="Lending")
                ws.cell(row=row, column=2, value=protocol_name.title())
                ws.cell(row=row, column=3, value=f"{protocol_name.title()} {item.get('underlying_token_name', '')}")
                ws.cell(row=row, column=4, value=round(float(item.get("balance", 0)), 8))
                ws.cell(row=row, column=5, value=f"${float(item.get('price', 0)):,.2f}")
                ws.cell(row=row, column=6, value=f"{float(item.get('apr', 0)):.2f}%")
                ws.cell(row=row, column=7, value=f"${float(item.get('usd_value', 0)):,.2f}")
                row += 1
        elif isinstance(protocol_data, dict) and "protocols" in protocol_data:
            # LayerBank format
            for sub_protocol_name, sub_protocol_data in protocol_data["protocols"].items():
                if "apr" in sub_protocol_data and "portfolio_entries" in sub_protocol_data["apr"]:
                    for entry in sub_protocol_data["apr"]["portfolio_entries"]:
                        # Get price from price data
                        price = 0
                        if "price" in sub_protocol_data and "token_prices" in sub_protocol_data["price"]:
                            price_data = sub_protocol_data["price"]["token_prices"].get(entry.get("explorer_address", "").lower(), {})
                            price = price_data.get("price", 0)
                        
                        # Get balance from token balances
                        balance = 0
                        token_balances = data.get("token_balances", [])
                        for token_data in token_balances:
                            if token_data["token"].get("address_hash", "").lower() == entry.get("explorer_address", "").lower():
                                balance = float(token_data["value"]) / (10 ** int(token_data["token"].get("decimals", 18)))
                                break
                        
                        usd_value = balance * price
                        action = "LEND" if entry.get("total_apr", 0) >= 0 else "BORROW"
                        
                        ws.cell(row=row, column=1, value="Lending")
                        ws.cell(row=row, column=2, value=sub_protocol_name.title())
                        ws.cell(row=row, column=3, value=f"{sub_protocol_name.title()} {action}")
                        ws.cell(row=row, column=4, value=round(balance, 8))
                        ws.cell(row=row, column=5, value=f"${price:,.2f}" if price > 0 else "N/A")
                        ws.cell(row=row, column=6, value=f"{entry.get('total_apr', 0):.2f}%")
                        ws.cell(row=row, column=7, value=f"${usd_value:,.2f}" if usd_value > 0 else "N/A")
                        row += 1
    
    # NFTs
    nft_valuations = data.get("nft_valuations", [])
    for nft_data in nft_valuations:
        ws.cell(row=row, column=1, value="NFT")
        ws.cell(row=row, column=2, value="Uniswap")
        ws.cell(row=row, column=3, value=nft_data.get("name", f"NFT #{nft_data.get('nft_id', '')}"))
        ws.cell(row=row, column=4, value="1")
        ws.cell(row=row, column=5, value="N/A")
        ws.cell(row=row, column=6, value="N/A")
        ws.cell(row=row, column=7, value=f"${float(nft_data.get('total_value_usd', 0)):,.2f}")
        row += 1
    
    # Merkle rewards
    merkle_rewards = data.get("merkle_rewards", [])
    if isinstance(merkle_rewards, dict):
        merkle_rewards = merkle_rewards.get("rewards", [])
    
    for reward_data in merkle_rewards:
        ws.cell(row=row, column=1, value="Reward")
        ws.cell(row=row, column=2, value="Merkle")
        ws.cell(row=row, column=3, value=f"Merkle Rewards ({reward_data.get('token', {}).get('symbol', '')})")
        ws.cell(row=row, column=4, value=reward_data.get("amount_formatted", "0"))
        ws.cell(row=row, column=5, value=f"${float(reward_data.get('token', {}).get('price', 0)):,.2f}")
        ws.cell(row=row, column=6, value="N/A")
        ws.cell(row=row, column=7, value=f"${float(reward_data.get('usd_value', 0)):,.2f}")
        row += 1
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

def create_summary_sheet(wb, address, data):
    """Create Summary sheet with totals and overview"""
    ws = wb.create_sheet("Summary")
    
    # Title
    ws.cell(row=1, column=1, value=f"Portfolio Summary - {address}")
    ws.cell(row=1, column=1).font = Font(bold=True, size=16)
    
    # Summary data
    row = 3
    
    # Extract data from the same structure as frontend
    token_balances = data.get("token_balances", [])
    nft_valuations = data.get("nft_valuations", [])
    merkle_rewards = data.get("merkle_rewards", [])
    if isinstance(merkle_rewards, dict):
        merkle_rewards = merkle_rewards.get("rewards", [])
    
    yield_tokens = data.get("yield_tokens", [])
    if isinstance(yield_tokens, dict):
        yield_tokens = yield_tokens.get("yield_tokens", [])
    
    lending_portfolio = data.get("lending_portfolio", {})
    
    # Calculate totals
    total_tokens = len(token_balances)
    total_value = sum(
        float(t["value"]) / (10 ** int(t["token"].get("decimals", 18))) * 
        (float(t["token"].get("exchange_rate", 0)) if t["token"].get("exchange_rate") else 0) 
        for t in token_balances
    )
    nft_count = len(nft_valuations)
    nft_value = sum(float(nft.get("total_value_usd", 0)) for nft in nft_valuations)
    reward_count = len(merkle_rewards)
    reward_value = sum(float(reward.get("usd_value", 0)) for reward in merkle_rewards)
    
    # Calculate yield token value
    yield_value = sum(float(yield_token.get("usd_value", 0)) for yield_token in yield_tokens)
    
    # Calculate lending value generically
    lending_value = 0
    for protocol_name, protocol_data in lending_portfolio.items():
        if isinstance(protocol_data, dict) and "portfolio_items" in protocol_data:
            # Tropykus format
            lending_value += sum(float(item.get("usd_value", 0)) for item in protocol_data["portfolio_items"])
        elif isinstance(protocol_data, dict) and "protocols" in protocol_data:
            # LayerBank format
            for sub_protocol_name, sub_protocol_data in protocol_data["protocols"].items():
                if "apr" in sub_protocol_data and "portfolio_entries" in sub_protocol_data["apr"]:
                    for entry in sub_protocol_data["apr"]["portfolio_entries"]:
                        # Get price and balance
                        price = 0
                        if "price" in sub_protocol_data and "token_prices" in sub_protocol_data["price"]:
                            price_data = sub_protocol_data["price"]["token_prices"].get(entry.get("explorer_address", "").lower(), {})
                            price = price_data.get("price", 0)
                        
                        balance = 0
                        for token_data in token_balances:
                            if token_data["token"].get("address_hash", "").lower() == entry.get("explorer_address", "").lower():
                                balance = float(token_data["value"]) / (10 ** int(token_data["token"].get("decimals", 18)))
                                break
                        
                        lending_value += balance * price
    
    summary_data = [
        ("Total Tokens", total_tokens),
        ("Total Token Value", f"${total_value:,.2f}"),
        ("NFT Count", nft_count),
        ("NFT Value", f"${nft_value:,.2f}"),
        ("Reward Count", reward_count),
        ("Reward Value", f"${reward_value:,.2f}"),
        ("Yield Token Value", f"${yield_value:,.2f}"),
        ("Lending Value", f"${lending_value:,.2f}"),
        ("Total Portfolio Value", f"${total_value + nft_value + reward_value + yield_value + lending_value:,.2f}"),
        ("Export Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    ]
    
    for label, value in summary_data:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=2, value=value)
        row += 1
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 8001))
    debug = os.environ.get("FLASK_ENV") == "development"
    app.run(host="0.0.0.0", port=port, debug=debug)